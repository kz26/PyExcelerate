# -*- coding: utf-8 -*-

from ..Workbook import Workbook
from ..Color import Color
from ..Font import Font
from ..Fill import Fill
from ..Style import Style
from ..Alignment import Alignment
import time
from datetime import datetime
import nose
from nose.tools import eq_, ok_, raises
from .utils import get_output_path

def test_style():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws[1][1].value = 1
	ws[1][2].value = 1
	ws[1][3].value = 1
	ws[1][1].style.font.bold = True
	ws[1][2].style.font.italic = True
	ws[1][3].style.font.underline = True
	ws[1][1].style.font.strikethrough = True
	ws[1][1].style.font.color = Color(255, 0, 255)
	ws[1][1].style.fill.background = Color(0, 255, 0)
	ws[1][2].style.fill.background = Color(255, 255, 0)
	ws[2][1].value = "asdf"
	ws.range("A2", "B2").merge()
	eq_(ws[1][2].value, ws[1][1].value)
	ws[2][2].value = "qwer"
	eq_(ws[1][2].value, ws[1][1].value)
	ws[2][1].style.fill.background = Color(0, 255, 0)
	ws[1][1].style.alignment.vertical = 'top'
	ws[1][1].style.alignment.horizontal = 'right'
	ws[1][1].style.alignment.rotation = 90
	eq_(ws[1][1].style.alignment.rotation, 90)
	ws[3][3].style.borders.top.color = Color(255, 0, 0)
	ws[3][3].style.borders.left.color = Color(0, 255, 0)
	ws[3][4].style.borders.right.style = '-.'
	ws[4][1].value = 'Lorem ipsum dolor sit amet, consectetur adipiscing elit. Nulla lobortis fermentum metus id congue. Sed ultrices velit id sapien sodales bibendum. Mauris volutpat porta arcu non bibendum. Pellentesque adipiscing lacus quam, ac congue ipsum fringilla sed. Praesent dapibus dignissim elit vel dictum. Pellentesque commodo iaculis ipsum a rhoncus. Sed mattis neque eget justo dignissim scelerisque. Nam odio neque, mattis et libero id, posuere aliquam mi.'
	ws[4][2].value = 'Lorem ipsum dolor sit amet, consectetur adipiscing elit. Nulla lobortis fermentum metus id congue. Sed ultrices velit id sapien sodales bibendum. Mauris volutpat porta arcu non bibendum. Pellentesque adipiscing lacus quam, ac congue ipsum fringilla sed. Praesent dapibus dignissim elit vel dictum. Pellentesque commodo iaculis ipsum a rhoncus. Sed mattis neque eget justo dignissim scelerisque. Nam odio neque, mattis et libero id, posuere aliquam mi.'
	ws[4][1].style.alignment.wrap_text = True
	wb.save(get_output_path("style-test.xlsx"))

@raises(TypeError)
def test_invalid_wrap_text():
	a = Alignment()
	a.wrap_text = True
	ok_(a.wrap_text)
	a.wrap_text = 'some random nonsense'

@raises(ValueError)
def test_invalid_horizontal():
	a = Alignment()
	a.horizontal = 'left'
	eq_(a.horizontal, 'left')
	a.horizontal = 'nowhere'

@raises(ValueError)
def test_invalid_vertical():
	a = Alignment()
	a.vertical = 'bottom'
	eq_(a.vertical, 'bottom')
	a.vertical = 'somewhere'

def test_style_compression():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws.range("A1","C3").value = 1
	ws.range("A1","C1").style.font.bold = True
	ws.range("A2","C3").style.font.italic = True
	ws.range("A3","C3").style.fill.background = Color(255, 0, 0)
	ws.range("C1","C3").style.font.strikethrough = True
	wb.save(get_output_path("style-compression-test.xlsx"))
	
def test_style_reference():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws[1][1].value = 1
	font = Font(bold=True, italic=True, underline=True, strikethrough=True)
	ws[1][1].style.font = font
	wb.save(get_output_path("style-reference-test.xlsx"))

def test_style_row():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws[1].style.fill.background = Color(255, 0, 0)
	ws[1][3].style.fill.background = Color(0, 255, 0)
	wb.save(get_output_path("style-row-test.xlsx"))

def test_style_row_col():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws[1][1].value = 'asdf'
	ws[1][3].value = 'asdf\nasdf\nasdf\nasdf'
	ws[3][1].value = 'asdfasdfasdfasdfasdfasdfasdfasdf'
	eq_(Style(), ws.get_row_style(1))
	eq_(Style(), ws.get_col_style(1))
	ws.set_row_style(1, Style(size=-1))
	ws.set_row_style(2, Style(size=-1))
	ws.set_row_style(3, Style(size=-1))
	ws.set_row_style(4, Style(size=0))
	ws.set_row_style(5, Style(size=100, fill=Fill(background=Color(0, 255, 0, 0))))
	ws.set_col_style(1, Style(size=-1))
	ws.set_col_style(2, Style(size=-1))
	ws.set_col_style(3, Style(size=-1))
	ws.set_col_style(4, Style(size=0))
	ws.set_col_style(5, Style(size=100, fill=Fill(background=Color(255, 0, 0, 0))))
	wb.save(get_output_path("style-auto-row-col-test.xlsx"))

def test_and_or_xor():
	bolditalic = Font(bold=True, italic=True)
	italicunderline = Font(italic=True, underline=True)
	eq_(Font(italic=True), bolditalic & italicunderline)
	eq_(Font(bold=True, italic=True, underline=True), bolditalic | italicunderline)
	eq_(Font(bold=True, underline=True), bolditalic ^ italicunderline)
	
	fontstyle = Style(font=Font(bold=True))
	fillstyle = Style(fill=Fill(background=Color(255, 0, 0, 0)))
	eq_(Style(), fontstyle & fillstyle)
	eq_(Style(font=Font(bold=True), fill=Fill(background=Color(255, 0, 0, 0))), fontstyle | fillstyle)
	eq_(Style(font=Font(bold=True), fill=Fill(background=Color(255, 0, 0, 0))), fontstyle ^ fillstyle)
	
	leftstyle = Style(alignment=Alignment('right', 'top'))
	bottomstyle = Style(alignment=Alignment(vertical='top', rotation=15))
	eq_(Style(alignment=Alignment('right', 'top', 15)), leftstyle | bottomstyle)
	eq_(Style(alignment=Alignment(vertical='top')), leftstyle & bottomstyle)
	eq_(Style(alignment=Alignment('right', rotation=15)), leftstyle ^ bottomstyle)
	
def test_str_():
	font = Font(bold=True, italic=True, underline=True, strikethrough=True)
	eq_(font.__repr__(), "<Font: Calibri, 11pt b i u s>")

def test_no_style_xml():
	try:
		import openpyxl
	except ImportError:
		raise nose.SkipTest('openpyxl not installed')
	data = [[1, 2, 3], [4, 5, 6], [7, 8, 9]] # data is a 2D array
	filename = get_output_path("no_style.xlsx")
	sheetname = "test"
	wb = Workbook()
	wb.new_sheet(sheetname, data=data)
	wb.save(filename)
	wbr = openpyxl.reader.excel.load_workbook(filename=filename,use_iterators=True)
	mySheet = wbr.get_sheet_by_name(sheetname)

def test_dense_sparse_styles():
	testData = [
		['1x1', '1x2', '1x3'],
		['2x1', '2x2', '2x3'],
		['3x1', '3x2', '3x3']
	]
	wb = Workbook()
	ws = wb.new_sheet("Test 1", data=testData)
	ws[2][2].style.font.bold = True
	ws[2][5] = '2x5'
	ws[2][7] = '2x7'
	ws[2][5].style.font.bold = True
	ws[5][2] = '5x2'
	ws[7][2] = '7x2'
	ws[5][2].style.font.bold = True
	ws[5][5] = '5x5'
	ws[5][5].style.font.bold = True
	ws[6][6] = '6x6'
	wb.save(get_output_path("dense-sparse-style-test.xlsx"))
	
def test_unicode_with_styles():
	wb = Workbook()
	ws = wb.new_sheet(u"ʇǝǝɥsǝpoɔıun")
	ws[1][1].value = u'Körperschaft des öffentlichen'
	ws.set_col_style(2, Style(size=0))
	wb.save(get_output_path("unicode-styles.xlsx"))
