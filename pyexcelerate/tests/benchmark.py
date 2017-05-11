from ..Workbook import Workbook
from ..Color import Color
from ..Style import Style
from ..Font import Font
from ..Fill import Fill
import time
from .utils import get_output_path
from random import randint
from concurrent.futures import ProcessPoolExecutor
from memory_profiler import memory_usage
import openpyxl
import xlsxwriter.workbook
import os
import gc

ROWS = 1000
COLUMNS = 100
BOLD = 1
ITALIC = 2
UNDERLINE = 4
RED_BG = 8
testData = [[1] * COLUMNS] * ROWS
formatData = [[1] * COLUMNS] * ROWS

def generate_data():
	global testData, formatData
	testData = [[1] * COLUMNS] * ROWS
	formatData = [[1] * COLUMNS] * ROWS
	for row in range(ROWS):
		for col in range(COLUMNS):
			formatData[row][col] = randint(1, (1 << 4) - 1)
	
def run_pyexcelerate_value_fastest():
	file = get_output_path('test_pyexcelerate_value_fastest.xlsx')
	stime = time.clock()
	wb = Workbook()
	ws = wb.new_sheet('Test 1', data=testData)
	wb.save(file)
	elapsed = time.clock() - stime
	return elapsed, os.path.getsize(file)

def run_pyexcelerate_value_faster():
	file = get_output_path('test_pyexcelerate_value_faster.xlsx')
	stime = time.clock()
	wb = Workbook()
	ws = wb.new_sheet('Test 1')
	for row in range(ROWS):
		for col in range(COLUMNS):
			ws.set_cell_value(row + 1, col + 1, 1)
	wb.save(file)
	elapsed = time.clock() - stime
	return elapsed, os.path.getsize(file)
	

def run_pyexcelerate_value_fast():
	file = get_output_path('test_pyexcelerate_value_fast.xlsx')
	stime = time.clock()
	wb = Workbook()
	ws = wb.new_sheet('Test 1')
	for row in range(ROWS):
		for col in range(COLUMNS):
			ws[row + 1][col + 1].value = 1
	wb.save(file)
	elapsed = time.clock() - stime
	return elapsed, os.path.getsize(file)
	
def run_openpyxl():
	file = get_output_path('test_openpyxl.xlsx')
	stime = time.clock()
	wb = openpyxl.workbook.Workbook(write_only=True) 
	ws = wb.create_sheet()
	ws.title = 'Test 1'
	for row in testData:
		ws.append(row)
	wb.save(file)
	elapsed = time.clock() - stime
	return elapsed, os.path.getsize(file)

def run_xlsxwriter_value():
	file = get_output_path('test_xlsxwriter.xlsx')
	stime = time.clock()
	wb = xlsxwriter.workbook.Workbook(file, {'constant_memory': True})
	ws = wb.add_worksheet()
	for row in range(ROWS):
		for col in range(COLUMNS):
			ws.write_number(row, col, 1)
	wb.close()
	elapsed = time.clock() - stime
	return elapsed, os.path.getsize(file)
	
def run_pyexcelerate_style_fastest():
	file = get_output_path('test_pyexcelerate_style_fastest.xlsx')
	stime = time.clock()
	wb = Workbook()
	ws = wb.new_sheet('Test 1')
	bold = Style(font=Font(bold=True))
	italic = Style(font=Font(italic=True))
	underline = Style(font=Font(underline=True))
	red = Style(fill=Fill(background=Color(255,0,0,0)))
	for row in range(ROWS):
		for col in range(COLUMNS):
			ws.set_cell_value(row + 1, col + 1, 1)
			style = Style()
			if formatData[row][col] & BOLD:
				style.font.bold = True
			if formatData[row][col] & ITALIC:
				style.font.italic = True
			if formatData[row][col] & UNDERLINE:
				style.font.underline = True
			if formatData[row][col] & RED_BG:
				style.fill.background = Color(255, 0, 0)
			ws.set_cell_style(row + 1, col + 1, style)
	wb.save(file)
	elapsed = time.clock() - stime
	return elapsed, os.path.getsize(file)
	
def run_pyexcelerate_style_faster():
	file = get_output_path('test_pyexcelerate_style_faster.xlsx')
	stime = time.clock()
	wb = Workbook()
	ws = wb.new_sheet('Test 1')
	for row in range(ROWS):
		for col in range(COLUMNS):
			ws.set_cell_value(row + 1, col + 1, 1)
			if formatData[row][col] & BOLD:
				ws.get_cell_style(row + 1, col + 1).font.bold = True
			if formatData[row][col] & ITALIC:
				ws.get_cell_style(row + 1, col + 1).font.italic = True
			if formatData[row][col] & UNDERLINE:
				ws.get_cell_style(row + 1, col + 1).font.underline = True
			if formatData[row][col] & RED_BG:
				ws.get_cell_style(row + 1, col + 1).fill.background = Color(255, 0, 0)
	wb.save(file)
	elapsed = time.clock() - stime
	return elapsed, os.path.getsize(file)
	
def run_pyexcelerate_style_fast():
	file = get_output_path('test_pyexcelerate_style_fast.xlsx')
	stime = time.clock()
	wb = Workbook()
	ws = wb.new_sheet('Test 1')
	for row in range(ROWS):
		for col in range(COLUMNS):
			ws[row + 1][col + 1].value = 1
			if formatData[row][col] & BOLD:
				ws[row + 1][col + 1].style.font.bold = True
			if formatData[row][col] & ITALIC:
				ws[row + 1][col + 1].style.font.italic = True
			if formatData[row][col] & UNDERLINE:
				ws[row + 1][col + 1].style.font.underline = True
			if formatData[row][col] & RED_BG:
				ws[row + 1][col + 1].style.fill.background = Color(255, 0, 0, 0)
	wb.save(file)
	elapsed = time.clock() - stime
	return elapsed, os.path.getsize(file)
	
def run_pyexcelerate_style_cheating():
	file = get_output_path('test_pyexcelerate_style_fastest.xlsx')
	stime = time.clock()
	wb = Workbook()
	ws = wb.new_sheet('Test 1')

	cell_formats = []

	for i in range(16):
		cell_format = Style()
		if i & BOLD:
			cell_format.font.bold = True
		if i & ITALIC:
			cell_format.font.italic = True
		if i & UNDERLINE:
			cell_format.font.underline = True
		if i & RED_BG:
			cell_format.fill.background = Color(255, 0, 0)
		cell_formats.append(cell_format)

	for row in range(ROWS):
		for col in range(COLUMNS):
			ws.set_cell_value(row + 1, col + 1, 1)
			ws.set_cell_style(row + 1, col + 1, cell_formats[formatData[row][col]])
	wb.save(file)
	elapsed = time.clock() - stime
	return elapsed, os.path.getsize(file)

def run_xlsxwriter_style_cheating():
	file = get_output_path('test_xlsxwriter_style.xlsx')
	stime = time.clock()
	wb = xlsxwriter.workbook.Workbook(file, {'constant_memory': True})
	ws = wb.add_worksheet()
	
	cell_formats = []

	for i in range(16):
		cell_format = wb.add_format()
		if i & BOLD:
			cell_format.set_bold()
		if i & ITALIC:
			cell_format.set_italic()
		if i & UNDERLINE:
			cell_format.set_underline()
		if i & RED_BG:
			cell_format.set_bg_color('red')
		cell_formats.append(cell_format)

	for row in range(ROWS):
		for col in range(COLUMNS):
			ws.write_number(row, col, 1, cell_formats[formatData[row][col]]) 
	wb.close()
	elapsed = time.clock() - stime
	return elapsed, os.path.getsize(file)
	
def run_xlsxwriter_style():
	file = get_output_path('test_xlsxwriter_style.xlsx')
	stime = time.clock()
	wb = xlsxwriter.workbook.Workbook(file, {'constant_memory': True})
	ws = wb.add_worksheet()
	
	for row in range(ROWS):
		for col in range(COLUMNS):
			format = wb.add_format()
			if formatData[row][col] & BOLD:
				format.set_bold()
			if formatData[row][col] & ITALIC:
				format.set_italic()
			if formatData[row][col] & UNDERLINE:
				format.set_underline()
			if formatData[row][col] & RED_BG:
				format.set_bg_color('red')
			ws.write_number(row, col, 1, format) 
	wb.close()
	elapsed = time.clock() - stime
	return elapsed, os.path.getsize(file)

def run_openpyxl_optimization():
	file = get_output_path('test_openpyxl_opt.xlsx')
	stime = time.clock()
	wb = openpyxl.workbook.Workbook() 
	ws = wb.create_sheet()
	ws.title = 'Test 1'
	for col_idx in range(COLUMNS):
		col = openpyxl.utils.get_column_letter(col_idx + 1)
		for row in range(ROWS):
			ws['%s%s'%(col, row + 1)].value = 1
			if formatData[row][col_idx] & BOLD:
				ws['%s%s'%(col, row + 1)].font = openpyxl.styles.Font(bold=True)
			if formatData[row][col_idx] & ITALIC:
				ws['%s%s'%(col, row + 1)].font = openpyxl.styles.Font(italic=True)
			if formatData[row][col_idx] & UNDERLINE:
				ws['%s%s'%(col, row + 1)].font = openpyxl.styles.Font(underline='single')
			if formatData[row][col_idx] & RED_BG:
				ws['%s%s'%(col, row + 1)].fill = openpyxl.styles.PatternFill(fill_type=openpyxl.styles.fills.FILL_SOLID,
						start_color=openpyxl.styles.Color(openpyxl.styles.colors.RED),
						end_color=openpyxl.styles.Color(openpyxl.styles.colors.RED))
			ws['%s%s'%(col, row + 1)].value = 1
	wb.save(file)
	elapsed = time.clock() - stime
	return elapsed, os.path.getsize(file)

def timeout_five(self, fnc, *args, **kwargs):
		with ProcessPoolExecutor() as p:
				f = p.submit(fnc, *args, **kwargs)
				return f.result(timeout=5)

def print_header():
	print(("{:15}{:>20}{:>10}{:>17}{:>20}").format("%dx%d" % (ROWS, COLUMNS), "name", "time (s)", "file size (kB)", "memory usage (MB)"))

def run(f):
	gc.collect()
	time, filesize = f()
	profile = memory_usage((f, ), interval=0.1)
	memory = max(profile) - min(profile)
	print(("{:>35}{:10.3f}{:17}{:20.3f}").format(*(f.__name__, time, filesize / 1024, memory)))

def test_all():
	global ROWS, COLUMNS
	ROWS = 100
	COLUMNS = 100
	generate_data()
	print_header()
	run(run_pyexcelerate_value_fastest)
	run(run_pyexcelerate_value_faster)
	run(run_pyexcelerate_value_fast)
	run(run_xlsxwriter_value)
	run(run_openpyxl)
	run(run_pyexcelerate_style_fastest)
	run(run_pyexcelerate_style_faster)
	run(run_pyexcelerate_style_fast)
	run(run_xlsxwriter_style)
	run(run_openpyxl_optimization)
	ROWS = 100
	COLUMNS = 1000
	generate_data()
	print_header()
	run(run_pyexcelerate_value_fastest)
	run(run_pyexcelerate_value_faster)
	run(run_pyexcelerate_value_fast)
	run(run_xlsxwriter_value)
	run(run_openpyxl)
	run(run_pyexcelerate_style_fastest)
	run(run_pyexcelerate_style_faster)
	run(run_pyexcelerate_style_fast)
	run(run_xlsxwriter_style)
	run(run_openpyxl_optimization)
	ROWS = 1000
	COLUMNS = 1000
	generate_data()
	print_header()
	run(run_pyexcelerate_value_fastest)
	run(run_pyexcelerate_value_faster)
	run(run_pyexcelerate_value_fast)
	run(run_xlsxwriter_value)
	run(run_openpyxl)
	run(run_pyexcelerate_style_fastest)
	run(run_pyexcelerate_style_faster)
	run(run_pyexcelerate_style_fast)
	run(run_xlsxwriter_style)
	run(run_openpyxl_optimization)
	